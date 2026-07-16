from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
AREA_CSV = ROOT / "data" / "processed" / "mall_area_reference.csv"
AUDIT_CSV = ROOT / "data" / "processed" / "mall_area_verification_audit.csv"
SOURCE_XLSX = ROOT / "data" / "raw" / "GBA_GLA_финальная_проверка.xlsx"


NAME_MAP = {
    "Жар-Птица НН": "Жар-Птица",
    "Фантастика НН": "Фантастика",
    "Моремолл Сочи": "МореМолл",
    "Небо НН": "Небо",
    "Аура Новосибирск": "Аура",
    "Седьмое небо НН": "Седьмое небо",
    "Автозаводец НН": "Автозаводец",
    "Павелецкая Плаза Москва": "Павелецкий Плаза",
    "Океанис НН": "Океанис",
    "Золотая Миля НН": "Золотая Миля",
    "Ганза НН": "Ганза",
    "Индиго НН": "Индиго",
    "Крым НН": "Крым",
    "Счастье НН": "Счастье",
    "Парк Авеню НН": "Парк Авеню",
    "РИО НН": "РИО",
}


# Values below reconcile the supplied workbook with stronger primary sources.
VERIFIED_OVERRIDES = {
    "Алмаз Челябинск": {
        "gla": None,
        "gba": 222000,
        "source": "https://www.almaz.center/about/",
        "status": "официальная инфографика объекта публикует общую площадь 222 000 м²; GLA не опубликована",
        "reliability": "высокая",
        "reason": "Добавлена подтвержденная GBA 222 000 м²; GLA оставлена пустой.",
    },
    "МореМолл": {
        "gla": 83000,
        "gba": 169000,
        "source": "data/raw/MOREMOLL.pdf, стр. 4; https://www.moremall.ru/",
        "status": "официальная презентация ТПС Недвижимость: GLA 83 000 м², GBA 169 000 м²",
        "reliability": "высокая",
        "reason": "В PDF указаны обе метрики; GLA 83 100 из таблицы заменена на 83 000.",
    },
    "Океанис": {
        "gla": 83000,
        "gba": 106000,
        "source": "https://oceanis-mall.ru/rents",
        "status": "официальный сайт: GBA 106 000 м²; GLA рассчитана как 35 000 + 40 000 + 5 000 + 3 000 м²",
        "reliability": "высокая",
        "reason": "Принята методика всей коммерчески используемой площади комплекса.",
    },
    "Индиго": {
        "gla": 40000,
        "gba": 60000,
        "source": "https://indigo-trk.ru/assets/files/Презентация Индиго_август_22.pdf",
        "status": "официальная презентация объекта: GLA около 40 000 м², GBA около 60 000 м²",
        "reliability": "высокая",
        "reason": "Обе метрики опубликованы в презентации на официальном домене.",
    },
    "Ганза": {
        "gla": 18600,
        "gba": 20663,
        "source": "https://shopandmall.ru/torgovye-centry/ganza-nijniy_novgorod-ul_rodionova_165_13",
        "status": "профильный каталог ShopAndMall; официальная публикация площади не найдена",
        "reliability": "средняя",
        "reason": "Рабочие значения из профильного каталога сохранены с пониженной надежностью.",
    },
    "Галерея Санкт-Петербург": {
        "gla": 97000,
        "gba": 192000,
        "source": "https://amp.rbc.ru/regional/spb_sz/08/04/2026/69d651af9a794724e695be55; https://www.malls.ru/rus/malls/20344.shtml",
        "status": "значения совпадают в актуальной публикации РБК и профильном каталоге",
        "reliability": "средняя",
        "reason": "Официальный сайт площадь не публикует; значения подтверждаются несколькими вторичными источниками.",
    },
    "Горизонт Ростов": {
        "gla": 133948,
        "gba": 184878,
        "source": "https://www.gorizontmall.ru/about/",
        "status": "официальный сайт прямо публикует GLA и GBA",
        "reliability": "высокая",
        "reason": "Значения из таблицы подтверждены без изменений.",
    },
    "Галерея Краснодар": {
        "gla": 64000,
        "gba": 145000,
        "source": "https://www.galleryk.ru/upload/iblock/ab5/prezentatsiya_galereya_krasnodar_2023.pdf",
        "status": "официальная презентация объекта: GLA 64 000 м², GBA 145 000 м²",
        "reliability": "высокая",
        "reason": "GLA 65 000 из таблицы исправлена на 64 000 по официальной презентации.",
    },
    "Аура Ярославль": {
        "gla": 63000,
        "gba": 120000,
        "source": "https://www.yarcom.ru/news/v-centre-yaroslavlya-otkryli-krupneyshiy-trc-aura-3498; https://www.malls.ru/rus/malls/117876.shtml",
        "status": "публикация об открытии: GLA около 63 000 м², GBA 120 000 м²",
        "reliability": "средняя",
        "reason": "GLA 55 000 из таблицы не подтверждена и заменена на устойчивое значение около 63 000.",
    },
    "Кристалл Тюмень": {
        "gla": 75000,
        "gba": 100000,
        "source": "https://trc-kristall.ru/renter/; https://new-retail.ru/novosti/retail/tjumenskij_trc_kristall_gotovitsja_k_otkrytiju/",
        "status": "GBA подтверждена официальным сайтом; GLA опубликована профильным изданием",
        "reliability": "средняя",
        "reason": "GLA 76 000 из таблицы заменена на 75 000; GBA 100 000 подтверждена официально.",
    },
    "Галерея Новосибирск": {
        "gla": 54000,
        "gba": 134000,
        "source": "https://galereya-novosibirsk.ru/upload/iblock/696/gallery_novosibirsk_2024_2.pptx.pdf",
        "status": "официальная презентация объекта 2024: GLA 54 000 м², GBA 134 000 м²",
        "reliability": "высокая",
        "reason": "Значения 54 700 / 125 000 из таблицы исправлены по официальной презентации.",
    },
    "Павелецкий Плаза": {
        "gla": None,
        "gba": 33000,
        "source": "https://pavplaza.ru/about/center/",
        "status": "официальный сайт публикует общую площадь 33 000 м²; GLA не опубликована",
        "reliability": "высокая",
        "reason": "Добавлена только подтвержденная GBA; GLA оставлена пустой.",
    },
}


def clean_number(value):
    if value in (None, "") or pd.isna(value):
        return None
    return int(float(value))


def load_workbook_rows() -> dict[str, dict]:
    workbook = load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    sheet = workbook["Итоговая таблица"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        mall = NAME_MAP.get(row["Объект"], row["Объект"])
        rows[mall] = row
    return rows


def main() -> None:
    reference = pd.read_csv(AREA_CSV).fillna("")
    incoming = load_workbook_rows()
    audit_rows = []

    for index, row in reference.iterrows():
        mall = row["ТЦ/ТРК"]
        source_row = incoming.get(mall, {})
        old_gla = clean_number(row["GLA"])
        old_gba = clean_number(row["GBA"])
        incoming_gla = clean_number(source_row.get("GLA, м²"))
        incoming_gba = clean_number(source_row.get("GBA, м²"))
        override = VERIFIED_OVERRIDES.get(mall)

        if override:
            final_gla = override["gla"]
            final_gba = override["gba"]
            reference.at[index, "GLA"] = "" if final_gla is None else final_gla
            reference.at[index, "GBA"] = "" if final_gba is None else final_gba
            reference.at[index, "Источник площади"] = override["source"]
            reference.at[index, "Статус площади"] = override["status"]
            reference.at[index, "Надежность"] = override["reliability"]
            decision = "Принято с корректировкой" if (incoming_gla, incoming_gba) != (final_gla, final_gba) else "Принято"
            reason = override["reason"]
        else:
            final_gla, final_gba = old_gla, old_gba
            decision = "Сохранено текущее значение" if (old_gla or old_gba) else "Данных нет"
            reason = "Текущий справочник содержит не менее надежный источник; слабее подтвержденные значения не замещают его."

        audit_rows.append(
            {
                "ТЦ/ТРК": mall,
                "GLA в файле": incoming_gla or "",
                "GBA в файле": incoming_gba or "",
                "GLA до проверки": old_gla or "",
                "GBA до проверки": old_gba or "",
                "GLA принято": final_gla or "",
                "GBA принято": final_gba or "",
                "Решение": decision,
                "Обоснование": reason,
            }
        )

    reference.to_csv(AREA_CSV, index=False, encoding="utf-8")
    pd.DataFrame(audit_rows).to_csv(AUDIT_CSV, index=False, encoding="utf-8")
    print(f"updated: {AREA_CSV}")
    print(f"audit: {AUDIT_CSV}")


if __name__ == "__main__":
    main()
